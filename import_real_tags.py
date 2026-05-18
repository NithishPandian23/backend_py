"""
import_real_tags.py
====================
Reads the two real SCADA tag spreadsheets from a wind farm site and
produces a unified tag list compatible with the existing backend Tag model.

Inputs (place in same folder as this script, or change the paths below):
  - R18C20SP1_DAC_01_001_client_imptag.xlsx   (Acciona AW77/1500 DAC platform)
  - R18GS25_ING_04_001_client_imptag.xlsx     (Acciona AW3000 INGECON platform)

Outputs:
  - wind_turbine_tags.json   (full tag list, used by upload_tags.py)
  - wind_turbine_tags.csv    (same content, for inspection in Excel)
  - tag_summary.txt          (human-readable counts by category and model)

Run:
  python import_real_tags.py
"""
from __future__ import annotations
import csv
import json
import re
import sys
from pathlib import Path
from typing import Optional

try:
    from openpyxl import load_workbook
except ImportError:
    print("ERROR: openpyxl not installed. Run: pip install openpyxl", file=sys.stderr)
    sys.exit(1)


# ---------------------------------------------------------------------------
# Configuration — edit these paths if your files are elsewhere
# ---------------------------------------------------------------------------
SCRIPT_DIR = Path(__file__).parent

INPUT_FILES = [
    {
        "path": SCRIPT_DIR / "R18C20SP1_DAC_01_001_client_imptag.xlsx",
        "model": "R18C20_DAC",
        "model_short": "C20",
        "model_label": "Acciona AW77 (DAC, 1.5 MW class)",
    },
    {
        "path": SCRIPT_DIR / "R18GS25_ING_04_001_client_imptag.xlsx",
        "model": "R18GS25_ING",
        "model_short": "GS25",
        "model_label": "Acciona AW3000 (INGECON, 3 MW class)",
    },
]

OUTPUT_JSON    = SCRIPT_DIR / "wind_turbine_tags.json"
OUTPUT_CSV     = SCRIPT_DIR / "wind_turbine_tags.csv"
OUTPUT_SUMMARY = SCRIPT_DIR / "tag_summary.txt"


# ---------------------------------------------------------------------------
# Spanish subsystem prefix → frontend category enum
#
# The 16 frontend categories are:
#   General · Meteorological · Rotor · Gearbox · Generator · Converter ·
#   Transformer · Yaw · Hydraulic · Cooling · Brake · Lubrication ·
#   Tower · Grid · Safety · Performance
# ---------------------------------------------------------------------------
PREFIX_CATEGORY: dict[str, str] = {
    # Specific subsystems
    "Giro":       "Yaw",            # Yaw / orientation
    "Hidr":       "Hydraulic",      # Hidráulica
    "Mul":        "Gearbox",        # Multiplicadora = gearbox
    "Gen":        "Generator",      # Generator (NOT "Generales")
    "GENSET":     "Generator",      # Auxiliary genset
    "Conv":       "Converter",      # Power converter
    "Trf":        "Transformer",
    "Amb":        "Meteorological", # Ambient/meteo
    "Refr":       "Cooling",        # Refrigeración
    "Freno":      "Brake",
    "Rotor":      "Rotor",
    "Red":        "Grid",           # Grid
    "Switchgear": "Grid",
    "Top":        "Tower",          # Tower top
    "Gon":        "Tower",          # Góndola (nacelle, lives on tower)
    "Gond":       "Tower",

    # Performance & power-control buckets
    "Gro":   "Performance",  # Grupo potencia (power control)
    "Turb":  "Performance",  # Turbine general performance signals
    "Prod":  "Performance",  # Production
    "Disp":  "Performance",  # Available power

    # Safety / protection
    "SPPF":  "Safety",       # Safety protection function

    # Operational/control miscellany → General
    "Aux":       "General",
    "Varios":    "General",
    "Generales": "General",
    "Oper":      "General",
    "Ctr":       "General",
    "Reg":       "General",
    "VR":        "General",
    "SP":        "General",
    "Guys":      "General",
    "Gem":       "General",
    "Com":       "General",
    "UY":        "General",
    "PF":        "General",
}

# Keywords in the OPC name/description that signal a Lubrication tag.
# Used as a post-process to reclassify tags that otherwise land in
# Gearbox/Generator/General.
LUBRICATION_KEYWORDS = ("lubric", "engrase", "grease", "lub.")

# Keywords that mark a tag as "critical" (alarm/fault/trip-class signal).
# Both Spanish and English variants since descriptions are bilingual.
CRITICAL_KEYWORDS = (
    "fallo", "error", "alarma", "alarm", "disparo", "trip", "fault",
    "failure", "emergenc", "overspeed", "overload", "sobrecarga",
    "stop", "parada", "critic",
)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def normalize_header(header: tuple) -> dict[str, int]:
    """Return {column-name → column-index} from a header row.
    Some files have unnamed columns (e.g. C20 cols 2-3 are USERCODE/Area but
    unlabeled in the sheet header). We label them positionally too.
    """
    idx: dict[str, int] = {}
    for i, h in enumerate(header):
        if h is None:
            continue
        key = str(h).strip()
        idx[key] = i
    return idx


def parse_float(val) -> Optional[float]:
    if val is None or val == "":
        return None
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).strip().replace(",", ".")
    if not s:
        return None
    try:
        return float(s)
    except ValueError:
        return None


def get_subsystem_prefix(opc_name: str) -> str:
    """A1.Giro.ErrSensorGiro → 'Giro'; A1.Bloqueo → 'Bloqueo' (single-token tags)."""
    parts = opc_name.split(".")
    return parts[1] if len(parts) >= 2 else parts[0]


def map_category(prefix: str, description: str, opc_name: str) -> str:
    """Map a tag to one of the 16 frontend categories."""
    base = PREFIX_CATEGORY.get(prefix, "General")

    # Lubrication reclassification — sniff the description and the OPC name
    haystack = f"{opc_name} {description}".lower()
    if any(kw in haystack for kw in LUBRICATION_KEYWORDS):
        return "Lubrication"

    return base


def derive_is_critical(description: str, opc_name: str) -> bool:
    """A tag is 'critical' if its description names a fault/alarm/trip event."""
    haystack = f"{opc_name} {description}".lower()
    return any(kw in haystack for kw in CRITICAL_KEYWORDS)


def derive_data_type(tipo: Optional[str]) -> str:
    if not tipo:
        return "FLOAT"
    tipo_upper = str(tipo).strip().upper()
    if tipo_upper == "DIGITAL":
        return "BOOL"
    if tipo_upper == "FLOAT":
        return "FLOAT"
    return "FLOAT"


def derive_default_bounds(data_type: str, unit: str) -> tuple[float, float]:
    """Sensible defaults for tags whose MIN/MAX are blank in the source file."""
    if data_type == "BOOL":
        return (0.0, 1.0)
    unit_l = (unit or "").lower()
    # A handful of common units have well-known operating ranges
    if "ºc" in unit_l or "°c" in unit_l:
        return (-50.0, 200.0)
    if unit_l in ("%",):
        return (0.0, 100.0)
    if unit_l in ("rpm",):
        return (0.0, 30.0)
    if unit_l in ("m/s",):
        return (0.0, 50.0)
    if unit_l in ("bar",):
        return (0.0, 300.0)
    if unit_l in ("kw", "kva", "kvar"):
        return (-5000.0, 5000.0)
    if unit_l in ("v",):
        return (0.0, 1000.0)
    if unit_l in ("a",):
        return (0.0, 2000.0)
    if unit_l in ("hz",):
        return (45.0, 55.0)
    if unit_l in ("°",):
        return (-360.0, 360.0)
    return (-1.0e6, 1.0e6)


def derive_update_interval(category: str, data_type: str) -> float:
    """Faster updates for high-frequency operational signals."""
    if category in ("Generator", "Converter", "Grid", "Performance"):
        return 1.0
    if category in ("Meteorological", "Rotor"):
        return 2.0
    if category in ("Yaw", "Brake", "Hydraulic"):
        return 5.0
    return 10.0


# ---------------------------------------------------------------------------
# Read one file → list of tag dicts
# ---------------------------------------------------------------------------

def read_file(spec: dict) -> list[dict]:
    fp = spec["path"]
    if not fp.exists():
        raise FileNotFoundError(
            f"Missing input file: {fp}\n"
            f"Place '{fp.name}' next to this script and rerun."
        )

    print(f"Reading {fp.name} …")
    wb = load_workbook(str(fp), read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if not rows:
        return []

    cols = normalize_header(rows[0])

    def get(row, key, default=""):
        i = cols.get(key)
        if i is None or i >= len(row):
            return default
        v = row[i]
        if v is None:
            return default
        return v

    out: list[dict] = []
    seen_opc: set[str] = set()
    blank_count = 0
    dup_count = 0

    for row in rows[1:]:
        opc_name = get(row, "OPC NAME")
        if not opc_name:
            blank_count += 1
            continue
        opc_name = str(opc_name).strip()
        if not opc_name:
            blank_count += 1
            continue

        if opc_name in seen_opc:
            dup_count += 1
            continue
        seen_opc.add(opc_name)

        description = str(get(row, "DESCRIPTION") or get(row, "DESCRIPCIÓN") or "").strip()
        tipo = get(row, "TIPO")
        logica_inactive = str(get(row, "LOGICA1") or "").strip()
        logica_active   = str(get(row, "LOGICA2") or "").strip()
        unit = str(get(row, "UNIDADES") or "").strip()
        min_raw = parse_float(get(row, "MIN"))
        max_raw = parse_float(get(row, "MAX"))

        prefix = get_subsystem_prefix(opc_name)
        data_type = derive_data_type(tipo)
        category = map_category(prefix, description, opc_name)

        default_min, default_max = derive_default_bounds(data_type, unit)
        min_value = min_raw if min_raw is not None else default_min
        max_value = max_raw if max_raw is not None else default_max
        # Guard against malformed bounds
        if min_value > max_value:
            min_value, max_value = default_min, default_max

        is_critical = derive_is_critical(description, opc_name)
        update_interval_sec = derive_update_interval(category, data_type)

        out.append({
            "tag_name":             opc_name,
            "description":          description or opc_name,
            "category":             category,
            "subsystem_code":       prefix,
            "data_type":            data_type,
            "unit":                 unit,
            "min_value":            float(min_value),
            "max_value":            float(max_value),
            "update_interval_sec":  float(update_interval_sec),
            "is_critical":          is_critical,
            "current_value":        None,
            "last_updated":         None,
            "turbine_model":        spec["model"],
            "turbine_model_label":  spec["model_label"],
            "logic_active_label":   logica_active or None,
            "logic_inactive_label": logica_inactive or None,
        })

    print(f"  → {len(out)} tags  (skipped {blank_count} blank rows, {dup_count} duplicates within file)")
    return out


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main() -> int:
    print(f"\n{'='*60}")
    print("  IMPORT REAL SCADA TAGS")
    print(f"{'='*60}\n")

    all_tags: list[dict] = []
    per_file_counts: list[tuple[str, int]] = []

    for spec in INPUT_FILES:
        tags = read_file(spec)
        all_tags.extend(tags)
        per_file_counts.append((spec["model"], len(tags)))

    # Assign sequential tag_id
    for i, t in enumerate(all_tags, start=1):
        t["tag_id"] = i

    # Sort: by category (ordered) then by tag_name for predictable IDs
    CATEGORY_ORDER = [
        "General", "Meteorological", "Rotor", "Gearbox", "Generator", "Converter",
        "Transformer", "Yaw", "Hydraulic", "Cooling", "Brake", "Lubrication",
        "Tower", "Grid", "Safety", "Performance",
    ]
    cat_rank = {c: i for i, c in enumerate(CATEGORY_ORDER)}
    all_tags.sort(key=lambda t: (
        t["turbine_model"],
        cat_rank.get(t["category"], 999),
        t["tag_name"],
    ))
    for i, t in enumerate(all_tags, start=1):
        t["tag_id"] = i

    # ----- Write JSON -----
    print(f"\nWriting {OUTPUT_JSON.name} …")
    with open(OUTPUT_JSON, "w", encoding="utf-8") as f:
        json.dump(all_tags, f, indent=2, ensure_ascii=False)

    # ----- Write CSV -----
    print(f"Writing {OUTPUT_CSV.name} …")
    csv_columns = [
        "tag_id", "tag_name", "description", "category", "subsystem_code",
        "data_type", "unit", "min_value", "max_value", "update_interval_sec",
        "is_critical", "turbine_model", "logic_active_label", "logic_inactive_label",
    ]
    with open(OUTPUT_CSV, "w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=csv_columns)
        writer.writeheader()
        for t in all_tags:
            writer.writerow({k: t.get(k, "") for k in csv_columns})

    # ----- Summary file -----
    by_category: dict[str, int] = {}
    by_model_category: dict[str, dict[str, int]] = {}
    by_data_type: dict[str, int] = {}
    critical_count = 0
    for t in all_tags:
        cat = t["category"]
        mdl = t["turbine_model"]
        by_category[cat] = by_category.get(cat, 0) + 1
        by_model_category.setdefault(mdl, {}).setdefault(cat, 0)
        by_model_category[mdl][cat] += 1
        by_data_type[t["data_type"]] = by_data_type.get(t["data_type"], 0) + 1
        if t["is_critical"]:
            critical_count += 1

    lines = []
    lines.append(f"REAL SCADA TAG IMPORT — SUMMARY")
    lines.append("=" * 60)
    lines.append(f"Total tags:        {len(all_tags)}")
    lines.append(f"Critical signals:  {critical_count}")
    lines.append(f"Data types:        {by_data_type}")
    lines.append("")
    lines.append("Per-file counts:")
    for mdl, n in per_file_counts:
        lines.append(f"  {mdl:<16} {n} tags")
    lines.append("")
    lines.append("Per-category counts:")
    for cat in CATEGORY_ORDER:
        n = by_category.get(cat, 0)
        if n > 0:
            lines.append(f"  {cat:<16} {n}")
    lines.append("")
    lines.append("Per-model × per-category:")
    for mdl, cats in by_model_category.items():
        lines.append(f"  {mdl}:")
        for cat in CATEGORY_ORDER:
            n = cats.get(cat, 0)
            if n > 0:
                lines.append(f"    {cat:<16} {n}")
    lines.append("")
    summary = "\n".join(lines)
    OUTPUT_SUMMARY.write_text(summary, encoding="utf-8")

    print(f"Writing {OUTPUT_SUMMARY.name} …")
    print()
    print(summary)
    print()
    print(f"✓ Done. Now run `python upload_tags.py` to push to Firestore.")
    return 0


if __name__ == "__main__":
    sys.exit(main())